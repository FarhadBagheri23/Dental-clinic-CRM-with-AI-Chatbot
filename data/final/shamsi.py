import pandas as pd, jdatetime, datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
F='dental_clinic_bi.xlsx'
S={n:pd.read_excel(F,sheet_name=n) for n in pd.ExcelFile(F).sheet_names}
FA='۰۱۲۳۴۵۶۷۸۹'
def fa(s): return ''.join(FA[int(c)] if c.isdigit() else c for c in str(s))
_c={}
def conv(ts):
    if pd.isna(ts): return (None,None,None,None)
    g=pd.Timestamp(ts).date()
    if g in _c: return _c[g]
    j=jdatetime.date.fromgregorian(date=g)
    mn=j.j_months_fa[j.month-1]
    out=(fa(f"{j.year}/{j.month:02d}/{j.day:02d}"), fa(f"{j.year}/{j.month:02d}"),
         f"{mn} {fa(j.year)}", f"{j.year}{j.month:02d}")
    _c[g]=out; return out
def add(df,src,pfx,month=True):
    d=pd.to_datetime(df[src],errors='coerce')
    r=[conv(x) for x in d]
    df[pfx+'_shamsi']=[x[0] for x in r]
    if month:
        df[pfx+'_month_shamsi']=[x[1] for x in r]
        df[pfx+'_month_label']=[x[2] for x in r]
        df[pfx+'_month_sort']=[x[3] for x in r]
    return df

S['fact_sessions']=add(S['fact_sessions'],'session_date','session')
S['fact_appointments']=add(S['fact_appointments'],'scheduled_datetime','appt')
S['fact_invoices']=add(S['fact_invoices'],'issue_date','issue')
S['fact_payments']=add(S['fact_payments'],'payment_date','payment')
S['fact_consumable_usage']=add(S['fact_consumable_usage'],'usage_date','usage')
S['patients']=add(S['patients'],'registration_date','reg')
S['treatment_plans']=add(S['treatment_plans'],'start_date','start')
S['treatment_plans']=add(S['treatment_plans'],'estimated_end_date','end',month=False)
S['treatment_sessions']=add(S['treatment_sessions'],'session_date','session',month=False)
S['appointments']=add(S['appointments'],'scheduled_datetime','appt',month=False)
S['invoices']=add(S['invoices'],'issue_date','issue',month=False)
S['payments']=add(S['payments'],'payment_date','payment',month=False)

with pd.ExcelWriter(F,engine='openpyxl') as w:
    for n,df in S.items(): df.to_excel(w,sheet_name=n,index=False)
wb=load_workbook(F)
hf=Font(name='Arial',bold=True,color='FFFFFF'); fl=PatternFill('solid',fgColor='1F4E79')
for ws in wb.worksheets:
    for c in ws[1]:
        c.font=hf; c.fill=fl; c.alignment=Alignment(horizontal='center',vertical='center')
        ws.column_dimensions[c.column_letter].width=max(12,min(28,len(str(c.value))+6))
    ws.freeze_panes='A2'
wb.save(F)
fs=S['fact_sessions']
print("sample:"); print(fs[['session_day','session_shamsi','session_month_label','session_month_sort']].head(3).to_string(index=False))
print("\ntrend axis values:"); print(fs.groupby('session_month_sort')['session_month_label'].first().to_string())
print("\ncols added to fact_sessions:",[c for c in fs.columns if 'shamsi' in c or 'month_label' in c or 'month_sort' in c])
